import logging
from telegram import Update, ForceReply
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes
from datetime import datetime
import openpyxl
import os
import json

# --- Cấu hình ---
TELEGRAM_BOT_TOKEN = "7928945035:AAERniPUUPlGTA1fkzd-PHn1i00Z5xUJ2rQ"  # THAY THẾ BẰNG TOKEN CỦA BẠN
DATA_FILE = "expenses_data.json"
CURRENCY_SYMBOL = "EUR" # Ký hiệu tiền tệ
CURRENCY_NAME = "Euro"  # Tên tiền tệ

# --- Logging ---
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.INFO
)
logger = logging.getLogger(__name__)

# --- Hàm trợ giúp lưu và tải dữ liệu (giữ nguyên) ---
def load_expenses():
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
                for chat_id_str, user_expenses in data.items():
                    for i, expense in enumerate(user_expenses):
                        if isinstance(expense['date'], str):
                            try:
                                data[chat_id_str][i]['date'] = datetime.fromisoformat(expense['date'])
                            except ValueError:
                                try:
                                    data[chat_id_str][i]['date'] = datetime.strptime(expense['date'], '%Y-%m-%d %H:%M:%S.%f')
                                except ValueError:
                                    logger.warning(f"Không thể phân tích ngày: {expense['date']} cho chat_id {chat_id_str}")
                return data
        except json.JSONDecodeError:
            logger.error(f"Lỗi giải mã JSON từ file {DATA_FILE}. Tạo file mới.")
            return {}
        except Exception as e:
            logger.error(f"Lỗi khi tải dữ liệu: {e}")
            return {}
    return {}

def save_expenses(all_user_expenses):
    try:
        data_to_save = {}
        for chat_id, user_expenses in all_user_expenses.items():
            data_to_save[str(chat_id)] = []
            for expense in user_expenses:
                expense_copy = expense.copy()
                if isinstance(expense_copy['date'], datetime):
                    expense_copy['date'] = expense_copy['date'].isoformat()
                data_to_save[str(chat_id)].append(expense_copy)
        with open(DATA_FILE, 'w', encoding='utf-8') as f:
            json.dump(data_to_save, f, ensure_ascii=False, indent=4)
    except Exception as e:
        logger.error(f"Lỗi khi lưu dữ liệu: {e}")

all_expenses_data = load_expenses()

# --- Các hàm xử lý logic của Bot ---
async def send_start_message(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user = update.effective_user
    await update.message.reply_html(
        rf"Xin chào {user.mention_html()}! Tôi là bot quản lý chi tiêu của bạn."
        f"\n\n<b>Cách thêm chi tiêu (đơn vị {CURRENCY_NAME}):</b>"
        "\nNhập trực tiếp: <Mô tả chi tiêu> <Số tiền>"
        f"\nVí dụ: <code>An sang 5.5</code> (nghĩa là 5.50 {CURRENCY_SYMBOL})"
        "\n\n<b>Các lệnh khác:</b>"
        "\n<b>xem</b> - Xem các chi tiêu đã nhập"
        "\n<b>xuat</b> - Xuất bảng chi tiêu Excel"
        "\n<b>xoa</b> <số_thứ_tự> - Xóa chi tiêu (xem số thứ tự bằng lệnh 'xem')"
        "\n<b>help</b> - Xem lại hướng dẫn này"
        "\n\nBạn cũng có thể dùng lệnh /start để xem lại hướng dẫn này.",
    )

async def send_help_message(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await update.message.reply_html(
        f"<b>Cách thêm chi tiêu (đơn vị {CURRENCY_NAME}):</b>"
        "\nNhập trực tiếp: <Mô tả chi tiêu> <Số tiền>"
        f"\nVí dụ: <code>An sang 5.5</code> hoặc <code>Mua sach 35.99</code>"
        "\n\n<b>Các lệnh khác:</b>\n"
        "<b>xem</b> - Xem các chi tiêu đã nhập\n"
        "<b>xuat</b> - Xuất bảng chi tiêu Excel\n"
        "<b>xoa</b> <số_thứ_tự> - Xóa chi tiêu (xem số thứ tự bằng lệnh 'xem')\n"
        "<b>help</b> - Xem lại hướng dẫn này"
    )

async def process_add_expense(update: Update, context: ContextTypes.DEFAULT_TYPE, description: str, amount_val: float) -> None:
    chat_id = update.message.chat_id
    try:
        if amount_val <= 0:
            await update.message.reply_text("Số tiền phải lớn hơn 0.")
            return
        if not description.strip():
            await update.message.reply_text("Mô tả không được để trống.")
            return

        entry_date = datetime.now()
        expense_entry = {
            "date": entry_date,
            "amount": amount_val, # Lưu số float
            "description": description.strip()
        }

        if str(chat_id) not in all_expenses_data:
            all_expenses_data[str(chat_id)] = []
        
        all_expenses_data[str(chat_id)].append(expense_entry)
        save_expenses(all_expenses_data)

        # Hiển thị với 2 chữ số thập phân cho Euro
        await update.message.reply_text(
            f"Đã thêm: {amount_val:.2f} {CURRENCY_SYMBOL} - {description.strip()} ({entry_date.strftime('%d/%m/%Y %H:%M')})"
        )

    except Exception as e:
        logger.error(f"Lỗi khi thêm chi tiêu: {e}")
        await update.message.reply_text("Đã xảy ra lỗi khi thêm chi tiêu. Vui lòng thử lại.")

async def process_view_expenses(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.message.chat_id
    user_expenses = all_expenses_data.get(str(chat_id), [])
    if not user_expenses:
        await update.message.reply_text("Bạn chưa có chi tiêu nào.")
        return
    response_message = "Danh sách chi tiêu của bạn (gõ 'xoa <số>' để xóa):\n"
    for i, expense in enumerate(user_expenses):
        date_obj = expense['date']
        if isinstance(date_obj, str):
            try: date_obj = datetime.fromisoformat(date_obj)
            except ValueError: date_obj = datetime.strptime(date_obj, '%Y-%m-%d %H:%M:%S.%f')
        date_str = date_obj.strftime('%d/%m/%Y %H:%M') if isinstance(date_obj, datetime) else str(date_obj)
        # Hiển thị với 2 chữ số thập phân cho Euro
        response_message += f"{i+1}. {date_str} - {expense['amount']:.2f} {CURRENCY_SYMBOL} - {expense['description']}\n"
    
    if len(response_message) > 4096:
        parts = [response_message[i:i+4000] for i in range(0, len(response_message), 4000)]
        for part in parts: await update.message.reply_text(part)
    else: await update.message.reply_text(response_message)

async def process_delete_expense(update: Update, context: ContextTypes.DEFAULT_TYPE, args: list) -> None:
    chat_id = update.message.chat_id
    user_expenses = all_expenses_data.get(str(chat_id), [])
    if not user_expenses:
        await update.message.reply_text("Bạn chưa có chi tiêu nào để xóa.")
        return
    try:
        if not args:
            await update.message.reply_text("Vui lòng cung cấp số thứ tự. Ví dụ: xoa 1")
            return
        index_to_delete = int(args[0]) - 1
        if 0 <= index_to_delete < len(user_expenses):
            deleted_expense = user_expenses.pop(index_to_delete)
            save_expenses(all_expenses_data)
            date_obj = deleted_expense['date']
            if isinstance(date_obj, str): date_obj = datetime.fromisoformat(date_obj)
            date_str = date_obj.strftime('%d/%m/%Y') if isinstance(date_obj, datetime) else str(date_obj)
            # Hiển thị với 2 chữ số thập phân cho Euro
            await update.message.reply_text(
                f"Đã xóa: {date_str} - {deleted_expense['amount']:.2f} {CURRENCY_SYMBOL} - {deleted_expense['description']}"
            )
        else:
            await update.message.reply_text(f"Số thứ tự không hợp lệ (1 đến {len(user_expenses)}).")
    except ValueError: await update.message.reply_text("Số thứ tự không hợp lệ.")
    except Exception as e:
        logger.error(f"Lỗi khi xóa chi tiêu: {e}")
        await update.message.reply_text("Đã xảy ra lỗi khi xóa chi tiêu.")

async def process_export_expenses(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.message.chat_id
    user_expenses = all_expenses_data.get(str(chat_id), [])
    if not user_expenses:
        await update.message.reply_text("Bạn chưa có chi tiêu nào để xuất.")
        return
    await update.message.reply_text("Đang chuẩn bị file Excel, vui lòng đợi...")
    try:
        workbook = openpyxl.Workbook()
        if "Sheet" in workbook.sheetnames: workbook.remove(workbook["Sheet"])
        expenses_by_month_year = {}
        for expense in user_expenses:
            entry_date = expense['date']
            if not isinstance(entry_date, datetime):
                try: entry_date = datetime.fromisoformat(str(entry_date))
                except ValueError:
                    logger.warning(f"Bỏ qua mục có ngày không hợp lệ: {entry_date}")
                    continue
            month_year_key = entry_date.strftime("%Y-%m")
            if month_year_key not in expenses_by_month_year: expenses_by_month_year[month_year_key] = []
            expenses_by_month_year[month_year_key].append(expense)
        if not expenses_by_month_year:
            await update.message.reply_text("Không có dữ liệu hợp lệ để xuất.")
            return
        sorted_month_year_keys = sorted(expenses_by_month_year.keys())
        for month_year_key in sorted_month_year_keys:
            month_expenses = expenses_by_month_year[month_year_key]
            dt_obj = datetime.strptime(month_year_key, "%Y-%m")
            sheet_name = f"Tháng {dt_obj.strftime('%m-%Y')}"
            sheet = workbook.create_sheet(title=sheet_name)
            
            headers = ["Ngày giờ nhập", f"Số tiền ({CURRENCY_SYMBOL})", "Mô tả"] # Cập nhật tiêu đề cột
            sheet.append(headers)
            for col_num, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_num); cell.font = openpyxl.styles.Font(bold=True)
                sheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = len(header) + 5
            sheet.column_dimensions['A'].width = 20; sheet.column_dimensions['B'].width = 18; sheet.column_dimensions['C'].width = 40 # Điều chỉnh độ rộng cột B
            
            total_amount_month = 0
            for expense_item in month_expenses:
                entry_date = expense_item['date']
                if not isinstance(entry_date, datetime): entry_date = datetime.fromisoformat(str(entry_date))
                row_data = [
                    entry_date.strftime("%d/%m/%Y %H:%M:%S"),
                    expense_item["amount"], # Giữ nguyên số float
                    expense_item["description"]
                ]
                sheet.append(row_data)
                # Định dạng ô số tiền là số có 2 chữ số thập phân
                current_row_idx = sheet.max_row
                sheet.cell(row=current_row_idx, column=2).number_format = '0.00' # hoặc '#,##0.00' nếu muốn dấu phẩy hàng ngàn
                total_amount_month += expense_item["amount"]
            
            sheet.append([])
            total_row = ["Tổng cộng tháng:", total_amount_month, ""]
            sheet.append(total_row)
            
            last_row_idx = sheet.max_row
            sheet.cell(row=last_row_idx, column=1).font = openpyxl.styles.Font(bold=True)
            sheet.cell(row=last_row_idx, column=2).font = openpyxl.styles.Font(bold=True)
            sheet.cell(row=last_row_idx, column=2).number_format = '0.00' # Định dạng tổng tiền
        
        filename = f"chi_tieu_{chat_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        workbook.save(filename)
        with open(filename, "rb") as f: await context.bot.send_document(chat_id=chat_id, document=f)
        os.remove(filename)
    except Exception as e:
        logger.error(f"Lỗi khi xuất Excel: {e}", exc_info=True)
        await update.message.reply_text(f"Đã xảy ra lỗi khi tạo file Excel.")

# --- Hàm xử lý tin nhắn văn bản (logic phân tích lệnh giữ nguyên) ---
async def handle_text_message(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not update.message or not update.message.text: return

    message_text = update.message.text.strip()
    parts = message_text.split()
    if not parts: return

    command_candidate = parts[0].lower()
    args = parts[1:]

    logger.info(f"Nhận: '{message_text}' từ user {update.effective_user.id}")

    if command_candidate == "start":
        await send_start_message(update, context)
    elif command_candidate == "help":
        await send_help_message(update, context)
    elif command_candidate == "xem":
        await process_view_expenses(update, context)
    elif command_candidate == "xuat":
        await process_export_expenses(update, context)
    elif command_candidate == "xoa":
        await process_delete_expense(update, context, args)
    elif len(parts) >= 2:
        potential_amount_str = parts[-1].replace(',', '.') # Cho phép nhập dấu phẩy làm phân cách thập phân
        description_parts = parts[:-1]
        try:
            amount_val = float(potential_amount_str)
            if amount_val > 0:
                description = " ".join(description_parts)
                await process_add_expense(update, context, description, amount_val)
            else:
                await update.message.reply_html(
                    f"Lệnh không được hỗ trợ hoặc cú pháp thêm chi tiêu sai.\n"
                    f"Thêm chi tiêu: <Mô tả> <Số tiền {CURRENCY_SYMBOL}>\n"
                    f"Ví dụ: <code>An trua 7.5</code>\n"
                    f"Gõ <code>help</code> để xem hướng dẫn."
                )
        except ValueError:
            await update.message.reply_html(
                f"Lệnh không được hỗ trợ hoặc cú pháp thêm chi tiêu sai.\n"
                f"Thêm chi tiêu: <Mô tả> <Số tiền {CURRENCY_SYMBOL}>\n"
                f"Ví dụ: <code>An trua 7.5</code>\n"
                f"Gõ <code>help</code> để xem hướng dẫn."
            )
    else:
        if len(parts) == 1 and command_candidate not in ["start", "help", "xem", "xuat"]:
             await update.message.reply_html(
                f"Lệnh không được hỗ trợ hoặc cú pháp thêm chi tiêu sai.\n"
                f"Thêm chi tiêu: <Mô tả> <Số tiền {CURRENCY_SYMBOL}>\n"
                f"Ví dụ: <code>An trua 7.5</code>\n"
                f"Gõ <code>help</code> để xem hướng dẫn."
            )

def main() -> None:
    application = Application.builder().token(TELEGRAM_BOT_TOKEN).build()
    application.add_handler(CommandHandler("start", send_start_message))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text_message))
    logger.info("Bot đang chạy...")
    application.run_polling()

if __name__ == "__main__":
    main()